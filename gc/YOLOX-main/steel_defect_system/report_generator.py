import os
import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib.colors import HexColor, white, black
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class ReportGenerator:
    def __init__(self, output_folder='static/reports'):
        self.output_folder = output_folder
        os.makedirs(output_folder, exist_ok=True)
        self.primary_color = HexColor('#2c7da0')
        self.secondary_color = HexColor('#61a5c2')
        self.accent_color = HexColor('#ef4444')
        self.success_color = HexColor('#22c55e')
        self.bg_light = HexColor('#f8fafc')

    def generate_single_image_report(self, image_url, detections, original_filename, 
                                     result_image_path=None):
        """生成单张图片的检测报告（PDF）"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        pdf_filename = f"report_{timestamp}_{original_filename.rsplit('.', 1)[0]}.pdf"
        pdf_path = os.path.join(self.output_folder, pdf_filename)

        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=A4,
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=25*mm,
            bottomMargin=20*mm
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            textColor=self.primary_color,
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=self.primary_color,
            spaceBefore=20,
            spaceAfter=12,
            fontName='Helvetica-Bold'
        )

        subheading_style = ParagraphStyle(
            'SubHeading',
            parent=styles['Heading3'],
            fontSize=12,
            textColor=HexColor('#64748b'),
            spaceBefore=15,
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )

        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
            spaceBefore=6,
            spaceAfter=6
        )

        elements = []

        elements.append(Paragraph("NEU-DET 钢材表面缺陷检测报告", title_style))
        elements.append(Spacer(1, 10))

        info_data = [
            ['报告编号', f'RPT-{timestamp}'],
            ['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['检测文件', original_filename],
            ['检测结果', f'发现 {len(detections)} 个缺陷' if detections else '未检测到缺陷'],
            ['检测状态', '✓ 质量合格' if not detections else '⚠ 发现异常']
        ]

        info_table = Table(info_data, colWidths=[100, 300])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), self.bg_light),
            ('TEXTCOLOR', (0, 0), (-1, -1), black),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#e2e8f0')),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 20))

        if result_image_path and os.path.exists(result_image_path):
            elements.append(Paragraph("检测图像", heading_style))
            try:
                img = RLImage(result_image_path, width=400*mm, height=280*mm)
                img.hAlign = 'CENTER'
                elements.append(img)
            except:
                elements.append(Paragraph("（图像加载失败）", body_style))
            elements.append(Spacer(1, 15))

        if detections:
            elements.append(Paragraph("缺陷详情", heading_style))
            
            defect_headers = ['#', '缺陷类别', '置信度', '边界框坐标', '尺寸 (px)', '面积 (px²)']
            defect_data = [defect_headers]
            
            for idx, det in enumerate(detections, 1):
                bbox = det.get('bbox', [0, 0, 0, 0])
                width = bbox[2] - bbox[0]
                height = bbox[3] - bbox[1]
                area = width * height
                
                row = [
                    str(idx),
                    det.get('class', '-'),
                    f"{det.get('confidence', 0) * 100:.2f}%",
                    f"[{bbox[0]:.0f}, {bbox[1]:.0f}, {bbox[2]:.0f}, {bbox[3]:.0f}]",
                    f"{width:.0f} × {height:.0f}",
                    f"{area:.0f}"
                ]
                defect_data.append(row)

            defect_table = Table(defect_data, colWidths=[30, 70, 60, 120, 70, 60])
            defect_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.primary_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), white),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#e2e8f0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, self.bg_light]),
            ]))
            elements.append(defect_table)
            
            elements.append(Spacer(1, 20))
            
            stats_elements = self._generate_statistics_section(detections, heading_style, subheading_style, body_style)
            elements.extend(stats_elements)

        footer_text = f"<br/><br/>---<br/>本报告由 NEU-DET 智能检测系统自动生成<br/>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        elements.append(Paragraph(footer_text, body_style))

        doc.build(elements)
        return pdf_path, pdf_filename

    def generate_single_excel_report(self, detections, original_filename):
        """生成单张图片的Excel报告"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"excel_{timestamp}_{original_filename.rsplit('.', 1)[0]}.xlsx"
        excel_path = os.path.join(self.output_folder, excel_filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "检测结果"

        header_fill = PatternFill(start_color="2c7da0", end_color="2c7da0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alt_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='e2e8f0'),
            right=Side(style='thin', color='e2e8f0'),
            top=Side(style='thin', color='e2e8f0'),
            bottom=Side(style='thin', color='e2e8f0')
        )

        ws.merge_cells('A1:F1')
        ws['A1'] = f"NEU-DET 检测报告 - {original_filename}"
        ws['A1'].font = Font(bold=True, size=14, color="2c7da0")
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A3'] = "报告编号:"
        ws['B3'] = f"RPT-{timestamp}"
        ws['A4'] = "生成时间:"
        ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A5'] = "文件名:"
        ws['B5'] = original_filename
        ws['A6'] = "缺陷数量:"
        ws['B6'] = str(len(detections))

        for cell in ['A3', 'A4', 'A5', 'A6']:
            ws[cell].font = Font(bold=True)

        headers = ['序号', '缺陷类别', '置信度 (%)', 'X1', 'Y1', 'X2', 'Y2', '宽度', '高度', '面积 (px²)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for idx, det in enumerate(detections, 1):
            row = 8 + idx
            bbox = det.get('bbox', [0, 0, 0, 0])
            width = bbox[2] - bbox[0]
            height = bbox[3] - bbox[1]
            area = width * height
            
            data = [
                idx,
                det.get('class', '-'),
                round(det.get('confidence', 0) * 100, 2),
                round(bbox[0], 1),
                round(bbox[1], 1),
                round(bbox[2], 1),
                round(bbox[3], 1),
                round(width, 1),
                round(height, 1),
                round(area, 1)
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                if idx % 2 == 0:
                    cell.fill = alt_fill

        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 14

        wb.save(excel_path)
        return excel_path, excel_filename

    def generate_batch_pdf_report(self, results, summary, folder_name):
        """生成文件夹批量检测的PDF报告"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        pdf_filename = f"batch_report_{timestamp}.pdf"
        pdf_path = os.path.join(self.output_folder, pdf_filename)

        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=A4,
            rightMargin=15*mm,
            leftMargin=15*mm,
            topMargin=20*mm,
            bottomMargin=15*mm
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=22,
            textColor=self.primary_color,
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=self.primary_color,
            spaceBefore=15,
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )

        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['Normal'],
            fontSize=9,
            leading=12,
            spaceBefore=4,
            spaceAfter=4
        )

        elements = []

        elements.append(Paragraph("NEU-DET 批量检测报告", title_style))
        elements.append(Paragraph(f"文件夹: {folder_name}", ParagraphStyle(
            'Subtitle', parent=body_style, alignment=TA_CENTER, fontSize=11, textColor=HexColor('#64748b')
        )))
        elements.append(Spacer(1, 15))

        summary_data = [
            ['统计项目', '数值'],
            ['总图片数', str(summary.get('total_images', 0))],
            ['缺陷图片数', str(summary.get('defect_images', 0))],
            ['总缺陷数', str(summary.get('total_detections', 0))],
            ['缺陷率', f"{summary.get('defect_rate', 0):.1f}%"],
            ['正常图片数', str(summary.get('total_images', 0) - summary.get('defect_images', 0))],
            ['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]

        summary_table = Table(summary_data, colWidths=[150, 200])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.primary_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('BACKGROUND', (0, 1), (0, -1), self.bg_light),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#e2e8f0')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        elements.append(Paragraph("检测结果明细", heading_style))

        detail_headers = ['#', '文件名', '状态', '缺陷数', '最高置信度']
        detail_data = [detail_headers]

        for idx, result in enumerate(results, 1):
            filename = result.get('filename', '-')
            has_defects = result.get('has_defects', False)
            detection_count = result.get('detection_count', 0)
            
            status = "⚠ 异常" if has_defects else "✓ 正常"
            
            max_confidence = 0
            if has_defects:
                detections = result.get('detections', [])
                if detections:
                    max_confidence = max(d.get('confidence', 0) for d in detections) * 100
            
            row = [
                str(idx),
                filename[:40] + '...' if len(filename) > 40 else filename,
                status,
                str(detection_count),
                f"{max_confidence:.1f}%" if max_confidence > 0 else "-"
            ]
            detail_data.append(row)

        detail_table = Table(detail_data, colWidths=[30, 180, 60, 50, 70])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.primary_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, self.bg_light]),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(detail_table)

        defect_results = [r for r in results if r.get('has_defects')]
        if defect_results:
            elements.append(Spacer(1, 20))
            elements.append(Paragraph(f"缺陷详情 ({len(defect_results)} 个异常项)", heading_style))
            
            for result in defect_results[:10]:
                filename = result.get('filename', '-')
                detections = result.get('detections', [])
                
                elements.append(Paragraph(f"<b>▸ {filename}</b>", subheading_style))
                
                if detections:
                    det_headers = ['类别', '置信度', '边界框']
                    det_data = [detectors]
                    
                    for det in detections[:5]:
                        bbox = det.get('bbox', [0, 0, 0, 0])
                        det_data.append([
                            det.get('class', '-'),
                            f"{det.get('confidence', 0)*100:.1f}%",
                            f"[{bbox[0]:.0f},{bbox[1]:.0f},{bbox[2]:.0f},{bbox[3]:.0f}]"
                        ])
                    
                    mini_table = Table(det_data, colWidths=[80, 60, 140])
                    mini_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#fef2f2')),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.3, HexColor('#fecaca')),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ]))
                    elements.append(mini_table)
                    elements.append(Spacer(1, 8))

        footer_text = f"<br/>---<br/>本报告由 NEU-DET 智能检测系统自动生成<br/>共检测 {len(results)} 张图片 | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        elements.append(Paragraph(footer_text, body_style))

        doc.build(elements)
        return pdf_path, pdf_filename

    def generate_batch_excel_report(self, results, summary, folder_name):
        """生成文件夹批量检测的Excel报告"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"batch_excel_{timestamp}.xlsx"
        excel_path = os.path.join(self.output_folder, excel_filename)

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "统计摘要"

        header_fill = PatternFill(start_color="2c7da0", end_color="2c7da0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alt_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
        danger_fill = PatternFill(start_color="fef2f2", end_color="fef2f2", fill_type="solid")
        success_fill = PatternFill(start_color="f0fdf4", end_color="f0fdf4", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='e2e8f0'),
            right=Side(style='thin', color='e2e8f0'),
            top=Side(style='thin', color='e2e8f0'),
            bottom=Side(style='thin', color='e2e8f0')
        )

        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = f"NEU-DET 批量检测报告 - {folder_name}"
        ws_summary['A1'].font = Font(bold=True, size=16, color="2c7da0")
        ws_summary['A1'].alignment = Alignment(horizontal='center')

        summary_info = [
            ('A3', '报告编号:', 'B3', f"BATCH-{timestamp}"),
            ('A4', '生成时间:', 'B4', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('A5', '文件夹:', 'B5', folder_name),
            ('A7', '总图片数:', 'B7', str(summary.get('total_images', 0))),
            ('A8', '缺陷图片数:', 'B8', str(summary.get('defect_images', 0))),
            ('A9', '总缺陷数:', 'B9', str(summary.get('total_detections', 0))),
            ('A10', '缺陷率:', 'B10', f"{summary.get('defect_rate', 0):.1f}%"),
            ('A11', '正常图片数:', 'B11', str(summary.get('total_images', 0) - summary.get('defect_images', 0))),
        ]

        for label_cell, label, value_cell, value in summary_info:
            ws_summary[label_cell] = label
            ws_summary[label_cell].font = Font(bold=True)
            ws_summary[value_cell] = value

        ws_detail = wb.create_sheet("检测结果详情")

        headers = ['序号', '文件名', '相对路径', '检测状态', '缺陷数量', '最高置信度']
        for col, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for idx, result in enumerate(results, 1):
            row = idx + 1
            has_defects = result.get('has_defects', False)
            detection_count = result.get('detection_count', 0)
            
            status = "异常" if has_defects else "正常"
            
            max_confidence = 0
            if has_defects:
                detections = result.get('detections', [])
                if detections:
                    max_confidence = max(d.get('confidence', 0) for d in detections) * 100
            
            data = [
                idx,
                result.get('filename', '-'),
                result.get('relative_path', '-'),
                status,
                detection_count,
                round(max_confidence, 2) if max_confidence > 0 else None
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_detail.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='center' if col != 2 else 'left')
                cell.border = border
                
                if idx % 2 == 0:
                    cell.fill = alt_fill
                
                if col == 4:
                    if value == "异常":
                        cell.fill = danger_fill
                        cell.font = Font(color="dc2626", bold=True)
                    else:
                        cell.fill = success_fill
                        cell.font = Font(color="166534", bold=True)

        column_widths = [8, 35, 45, 12, 12, 14]
        for col, width in enumerate(column_widths, 1):
            ws_detail.column_dimensions[get_column_letter(col)].width = width

        defect_results = [r for r in results if r.get('has_defects')]
        if defect_results:
            ws_defects = wb.create_sheet("缺陷详情")
            
            defect_headers = ['序号', '文件名', '缺陷序号', '缺陷类别', '置信度', 
                             'X1', 'Y1', 'X2', 'Y2', '宽度', '高度', '面积']
            for col, header in enumerate(defect_headers, 1):
                cell = ws_defects.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            global_idx = 1
            for result in defect_results:
                filename = result.get('filename', '-')
                detections = result.get('detections', [])
                
                for det_idx, det in enumerate(detections, 1):
                    row = global_idx + 1
                    bbox = det.get('bbox', [0, 0, 0, 0])
                    width = bbox[2] - bbox[0]
                    height = bbox[3] - bbox[1]
                    area = width * height
                    
                    data = [
                        global_idx,
                        filename,
                        det_idx,
                        det.get('class', '-'),
                        round(det.get('confidence', 0) * 100, 2),
                        round(bbox[0], 1),
                        round(bbox[1], 1),
                        round(bbox[2], 1),
                        round(bbox[3], 1),
                        round(width, 1),
                        round(height, 1),
                        round(area, 1)
                    ]
                    
                    for col, value in enumerate(data, 1):
                        cell = ws_defects.cell(row=row, column=col, value=value)
                        cell.alignment = Alignment(horizontal='center' if col not in [2, 4] else 'left')
                        cell.border = border
                        
                        if global_idx % 2 == 0:
                            cell.fill = alt_fill
                    
                    global_idx += 1

            defect_widths = [8, 30, 10, 15, 10, 8, 8, 8, 8, 8, 8, 10]
            for col, width in enumerate(defect_widths, 1):
                ws_defects.column_dimensions[get_column_letter(col)].width = width

        wb.save(excel_path)
        return excel_path, excel_filename

    def _generate_statistics_section(self, detections, heading_style, subheading_style, body_style):
        """生成统计分析部分内容"""
        elements = []
        
        elements.append(Paragraph("统计分析", heading_style))
        
        total_detections = len(detections)
        avg_confidence = sum(d.get('confidence', 0) for d in detections) / total_detections if total_detections > 0 else 0
        max_confidence = max((d.get('confidence', 0) for d in detections), default=0)
        min_confidence = min((d.get('confidence', 0) for d in detections), default=0)
        
        class_counts = {}
        for det in detections:
            cls = det.get('class', 'unknown')
            class_counts[cls] = class_counts.get(cls, 0) + 1
        
        stats_data = [
            ['指标', '数值'],
            ['总缺陷数', str(total_detections)],
            ['平均置信度', f"{avg_confidence * 100:.2f}%"],
            ['最高置信度', f"{max_confidence * 100:.2f}%"],
            ['最低置信度', f"{min_confidence * 100:.2f}%"]
        ]
        
        stats_table = Table(stats_data, colWidths=[150, 150])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.secondary_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#e2e8f0')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 15))
        
        if class_counts:
            elements.append(Paragraph("缺陷类别分布", subheading_style))
            
            class_data = [['缺陷类别', '数量', '占比']]
            for cls, count in sorted(class_counts.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_detections) * 100
                class_data.append([cls, str(count), f"{percentage:.1f}%"])
            
            class_table = Table(class_data, colWidths=[120, 80, 80])
            class_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.accent_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#fecaca')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, HexColor('#fff5f5')]),
            ]))
            elements.append(class_table)
        
        return elements
